#!/usr/bin/env python
# -*-coding:utf-8 -*-
# @Time    : 2021/07/27 0:15
# @Author  : Joy
# @Version : python3.9
# @Desc    : Excel表格处理


import os
import openpyxl
import xlrd
import xlwt
from xlutils.copy import copy


def create_xlsx_with_head(file_path, sheet_name, head_values=None):
    """
    创建空Excel表格
    :param file_path: 指定文件路径
    :param sheet_name: sheet名
    :param head_values: 头列表
    :return:
    """
    if head_values is None:
        head_values = ["Title", "Title-CN", "Date", "URL", "Text", "Text-CN"]
    if file_path.endswith(".xlsx"):
        dir_name = os.path.dirname(file_path)
        if not os.path.isdir(dir_name):
            os.makedirs(dir_name)
        # 新建文件和表格
        wb = openpyxl.Workbook()
        # 删除默认生成的sheet
        wb.remove(wb["Sheet"])
        sheet = wb.create_sheet(sheet_name, index=1)
        # 调整列宽
        sheet.column_dimensions['A'].width = 20.0
        sheet.column_dimensions['B'].width = 20.0
        sheet.column_dimensions['E'].width = 20.0
        sheet.column_dimensions['F'].width = 20.0
        # 添加表头
        if head_values: sheet.append(head_values)
        wb.save(file_path)  # 保存文件，注意以xlsx为文件扩展名
    else:
        print("file path must end with .xlsx")


def write_xlsx_apend(file_path, values):
    """
    向表格中增加行
    :param file_path: 文件路径
    :param values: 对象
    :return:
    """
    if os.path.isfile(file_path) and file_path.endswith(".xlsx"):
        wb = openpyxl.load_workbook(file_path)
        # 获取workbook中第一个表格
        sheet = wb[wb.sheetnames[0]]
        for value in values:
            # TODO：需修改成对象属性或者其他
            row = [value.title, value.title_cn, value.date, value.url, value.text, value.text_cn]
            sheet.append(row)
        wb.save(file_path)  # 保存文件，注意以xlsx为文件扩展名
    else:
        print("file path must end with .xlsx")



def write_excel_xls(path, sheet_name, value):
    index = len(value)  # 获取需要写入数据的行数
    workbook = xlwt.Workbook()  # 新建一个工作簿
    sheet = workbook.add_sheet(sheet_name)  # 在工作簿中新建一个表格
    for i in range(0, index):
        for j in range(0, len(value[i])):
            sheet.write(i, j, value[i][j])  # 像表格中写入数据（对应的行和列）
    workbook.save(path)  # 保存工作簿
    print("xls格式表格【写入】数据成功！")


def write_excel_xls_append(path, value):
    index = len(value)  # 获取需要写入数据的行数
    if index <= 0: return
    workbook = xlrd.open_workbook(path)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
    rows_old = worksheet.nrows  # 获取表格中已存在的数据的行数
    new_workbook = copy(workbook)  # 将xlrd对象拷贝转化为xlwt对象
    new_worksheet = new_workbook.get_sheet(0)  # 获取转化后工作簿中的第一个表格
    for i in range(0, index):
        # 追加写入数据，注意是从i+rows_old行开始写入
        # 表头："Title", "Date", "URL", "Text"
        new_worksheet.write(i + rows_old, 0, value[i].title)
        new_worksheet.write(i + rows_old, 1, value[i].date)
        new_worksheet.write(i + rows_old, 2, value[i].url)
        new_worksheet.write(i + rows_old, 3, value[i].text)
    new_workbook.save(path)  # 保存工作簿
    print("xls格式表格【追加】数据成功！")


def read_excel_xls(path):
    workbook = xlrd.open_workbook(path)  # 打开工作簿
    sheets = workbook.sheet_names()  # 获取工作簿中的所有表格
    worksheet = workbook.sheet_by_name(sheets[0])  # 获取工作簿中所有表格中的的第一个表格
    for i in range(0, worksheet.nrows):
        for j in range(0, worksheet.ncols):
            print(worksheet.cell_value(i, j), "\t", end="")  # 逐行逐列读取数据
        print()